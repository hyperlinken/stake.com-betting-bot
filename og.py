import random
import time
import keyboard
from mouse_dragger import ControllableDragger, DragConfig
from pixeldetector import ScreenPixelDetector
import datetime
import pyautogui

from openpyxl import Workbook, load_workbook
import os
import pandas as pd

excel_file = "live_bet_log.xlsx"

# Create workbook and headers if file does not exist
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bet Log"
    ws.append(["Index", "Betting Amount", "Result", "State", "Timestamp", "Bets", "PnL", "BaseAmt"])
    wb.save(excel_file)

def log_bet_to_excel(index, betting_amount, result, state, timestamp, bets, pnl, baseamt):
    wb = load_workbook(excel_file)
    ws = wb.active
    ws.append([
        index,
        betting_amount,
        result,
        str(state),  # store state array as string
        timestamp,
        bets,
        pnl,
        baseamt
    ])
    wb.save(excel_file)




# Simulate 4 sets, each with 8 cycles
# for set_num in range(1, 5):
#     for cycle_num in range(1, 9):
#         # Simulate your actual data here
        

# # Convert to DataFrame
# df = pd.DataFrame(all_cycles)

# # Save to Excel
# df.to_excel('cycle_data.xlsx', index=False)

# def log_bet_to_excel(index, betting_amount, result, state, timestamp, bets, pnl, baseamt):
    # all_cycles = []

    # cycle_data = 
    # all_cycles.append(cycle_data)

    # ws.append([
    #     index,
    #     betting_amount,
    #     result,
    #     str(state),  # store state array as string
    #     timestamp,
    #     bets,
    #     pnl,
    #     baseamt
    # ])
    # df = pd.DataFrame({
    #         'Index': index,
    #         'BETTING_AMT': betting_amount,
    #         'RESULT': result,     
    #         'STATE': state,      
    #         'TIMESTAMP': timestamp,  
    #         'BETS': bets,        
    #         'PNL': pnl,         
    #         'BASEAMT': baseamt      
    #     })
    # df.to_excel('cycle_data.xlsx', index=False)


# Create instance with default config
mouse = ControllableDragger()

# Or with custom config
# config = DragConfig(
#     base_speed=0.1,
#     curvature=0.1 + random.random() * 0.1,
#     jitter=1.0
# )


config = DragConfig(
    base_speed=0.00000001,     # ULTRA small, faster than blink
    curvature=0,         # almost straight line
    jitter=0,            # very low random movement
    click_delay=0,      # tiny pause before clicking
    typing_speed=0.00000001,     # almost instant typing
    typing_variation=0, # tiny randomness in typing
    momentum=False,        # NO momentum = full speed always
    micro_pauses=False     # no unnecessary pauses
)




mouse = ControllableDragger(config)

# Use the dragger
# mouse.move_to(441, 457)
# mouse.move_click_type(441, 457, "sourabh sexy")
# mouse.move_and_click(sourabh sexy1000, 200)
# mouse.drag_to(1200, 600)

# arr[13] = [0]

indicator = ScreenPixelDetector()
count=0
# while( True ):
#     if keyboard.is_pressed('g'):
#         break
#     success = indicator.get_pixel_color(1404, 797)
#     print(f"Pixel color at (1267, 686): {success}")
#     count += 1
#     if(success == (0,228,73)):
#         break
#         # arr[count%13] = 1
#         # print(arr[count%13])
#     time.sleep(0.3)

# while(True):
#     turn1 = indicator.get_pixel_color(1267, 686)
#     turn2 = indicator.get_pixel_color(1282, 683)
#     turn3 = indicator.get_pixel_color(1275, 637)

#     if(turn1 == (241,2,96) or turn2 == (241,2,96) or turn3 == (241,2,96)):
#         count += 1
#         print(count)
#     time.sleep(0.3)
B=7
arr = [B,B,B,B,B,B,B,B,B,B,B,B,B,B,B,B,B,B,B,B]
originalarrbackup = arr.copy()

def wait_success():
    while(True):
        # success = indicator.get_pixel_color(1404, 797)
        success = indicator.detect_color_in_region(1321, 769, 50, 50, (0, 228, 73))
        if(not success):
            break

def wait_turn():
    while(True):
        # print("hi")

        # turn1 = indicator.get_pixel_color(1267, 686)
        # turn2 = indicator.get_pixel_color(1282, 683)
        # turn3 = indicator.get_pixel_color(1275, 637)
        # if(turn1 != (241,2,96)):
        #     if(turn2 != (241,2,96)):
        #         if(turn3 != (241,2,96)):
        #             break

        turn = indicator.detect_color_in_region(1163, 639, 200, 100, (241, 2, 96))
        if( not turn ):
            break

def movewrite( x , y , w):
    pyautogui.moveTo(x,y)
    pyautogui.click()
    pyautogui.hotkey('ctrl', 'a')
    pyautogui.press('backspace')
    pyautogui.write(w)

def movetoclick( x, y):
    pyautogui.moveTo(x,y)
    pyautogui.click()

def putBet(currentbet):
    global multiplier
    # mouse.move_and_click(466, 366)
    # mouse.move_click_type( 466, 366 , f"{amt}")                                     #sets value of bet

    # mouse.clear_and_type(466, 366, f"{round(amt , 5)}", enter=False)                        #clears value and sets value of bet 
    # if(currentbet < 3):
    #     mouse.clear_and_type(415, 453, f"{2}", enter=False)
    #     multiplier = 2
    # elif(currentbet < 8 and  currentbet >= 3):
    #     mouse.clear_and_type(415, 453, f"{1.5}", enter=False)
    #     multiplier = 1.5
    # else:
    #     mouse.clear_and_type(415, 453, f"{1.7}", enter=False)
    #     multiplier = 1.7
    
    movewrite( 466, 366, f"{round(amt , 5)}" )

    # if(currentbet == 0):
    #     multiplier = 2 + 1/16

    # if(currentbet < 8):
    #     multiplier = 2
    # else:
    #     multiplier = 1.5
    multiplier = 2

    movewrite(415, 453, f"{round(multiplier , 2)}")
    # if(currentbet < 3):
    #     movewrite(415, 453, f"{2}")
    #     multiplier = 2
    # elif(currentbet < 8 and  currentbet >= 3):
    #     movewrite(415, 453, f"{1.5}")
    #     multiplier = 1.5
    # elif(currentbet < 14 and  currentbet >= 8):
    #     movewrite(415, 453, f"{1.7}")
    #     multiplier = 1.7
    # else:
    #     movewrite(415, 453, f"{1.5}")
    #     multiplier = 1.5


    while(True):
        betting1 = indicator.detect_color_in_region(336, 590, 500, 50, (0, 230, 1))
        betting2 = indicator.detect_color_in_region(336, 590, 500, 50, (31, 255, 32))

        betting3 = indicator.detect_color_in_region(336, 590, 500, 50, (0, 230, 73))
        betting4 = indicator.detect_color_in_region(336, 590, 500, 50, (26, 44, 56))

        if(betting1 or betting2 or betting3 or betting4):
            # mouse.move_and_click(603 , 608)                                         #sets bet
            movetoclick(603 , 608)          # it is same thing of upper line
            # mouse.move_to(603 , 608)                                              #trial
            # mouse.move_and_click(466, 366)
            # movetoclick(466 , 366)
            break
     
pnl = 0
baseamt = 1.7
# fixvalue = baseamt
# amt = baseamt/(2**13)
div = 8
amt = baseamt/(2**div)
bets = 0
# currentbet =0

while(True):
    
    putBet(count%div)

    print(count , "betting amount" , amt)

    while(True):
        betconfirm = indicator.get_pixel_color(380, 714)
        turn = indicator.detect_color_in_region(1163, 639, 200, 100, (241, 2, 96))

        if(betconfirm == (47, 69, 83)):
            flag = 1
            break
        if(turn):
            flag = 0
            break   

    if(flag == 0):
        continue

    if(flag == 1):
        currentbet = count%div
        # print(count , "betting amount" , amt)
        while(True):
            turn = indicator.detect_color_in_region(1163, 639, 200, 100, (241, 2, 96))
            # turn1 = indicator.get_pixel_color(1267, 686)
            # turn2 = indicator.get_pixel_color(1282, 683)
            # turn3 = indicator.get_pixel_color(1275, 637)
            success = indicator.detect_color_in_region(1321, 769, 50, 50, (0, 228, 73))
            # avg_color = inidicator.get_region_color(747, 242, 1000, 40, 'average')

            if(success):  #for 1.5x indicator
                arr[currentbet] = 1
                result = "success"
                print("success")
                print(arr)
                pnl += (amt * multiplier) - amt
                baseamt += (amt * multiplier) - amt
                break
            # if(turn1 == (241,2,96) or turn2 == (241,2,96) or turn3 == (241,2,96)): #for next bet 
            if(turn):
                arr[currentbet] = 0
                result = "fail"
                print("fail")
                print(arr)
                pnl -= amt
                baseamt -= amt               
                break

        # if(currentbet < 2):
        #     amt = 0        #amt double
        # elif(currentbet == 2):
        #     amt = baseamt/(2**13)
        # else:
        amt *= 2
        
        if(baseamt <= 0):
            print("chud gaye guru in bets : " , bets)
            break
        
        state = arr

        # if(currentbet < 3):
        #         paster = arr[0]
        #         past = arr[1]
        #         present = arr[2]
        #         if( paster == 1 and past == 1 and present == 1):     #base value reset and array reset
        #             arr.clear()
        #             arr.extend(originalarrbackup)
        #             amt = baseamt/(2**13)          #accordig to fix value reset
                    # mouse.move_to(470,164)
                    # mouse.move_and_click(470, 364)
                    # count = -1                       #count reset to -1
        if(currentbet >= 0):
            for i in range(1 , currentbet + 1):
                past = arr[i-1]
                present = arr[i]
                if( past == 1 and present == 1):     #base value reset and array reset
                    arr.clear()
                    arr.extend(originalarrbackup)
                    amt = baseamt/(2**16)          #accordig to fix value reset
                    # amt = 0
                    # mouse.move_to(470,164)
                    # mouse.move_and_click(470, 364)
                    count = -1                       #count reset to -1
                    break
        # elif(currentbet >= 7):
        #     arr.clear()
        #     arr.extend(originalarrbackup)
        #     amt = baseamt/(2**13)          #accordig to fix value reset
        #     # mouse.move_to(470,164)
        #     # mouse.move_and_click(470, 364)
        #     count = -1  ######################################
        # elif(currentbet >= 8):
        #     for i in range(8 , len(arr)):       
        #         present = arr[i]
        #         if( present == 1):                  #base value reset and array reset
        #             arr.clear()
        #             arr.extend(originalarrbackup)
        #             amt = baseamt/(2**13)          #accordig to fix value reset
        #             # mouse.move_to(470,164)
        #             # mouse.move_and_click(470, 364)
        #             count = -1                       #count reset
        #             break
                             #count reset
        # if(currentbet == 0):
        #     if(arr[0] == 1):
        #         arr.clear()
        #         arr.extend(originalarrbackup)
        #         amt = baseamt/(2**14)
        #         count = -1
        # else:
        #     for i in range(1 , currentbet + 1):
        #         past = arr[i-1]
        #         present = arr[i]
        #         if( past == 1 and present == 1):     #base value reset and array reset
        #             arr.clear()
        #             arr.extend(originalarrbackup)
        #             amt = baseamt/(2**14)
        #             count = -1
        #             break
        
        print(datetime.datetime.now().time() , "    bets:" , bets , "pnl:" , pnl , "baseamt:" , baseamt, "\n")

        log_bet_to_excel(
            count,
            amt,
            result,
            state,
            datetime.datetime.now().time(),
            bets,
            pnl,
            baseamt
        )
        count += 1
        bets += 1
        wait_success()
    
    wait_turn()



# update this code without altering the algorithm such that changes reflect in real time as terminal